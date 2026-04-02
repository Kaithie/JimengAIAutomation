# -*- coding: utf-8 -*-
"""
主窗口模块
应用程序的主界面
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import List, Dict, Optional, Callable
from pathlib import Path
import threading
import time
import os

from .settings_dialog import SettingsDialog
from .segment_table import SegmentTable
from .material_panel import MaterialPanel
from core.config import ConfigManager
from core.ai_engine import AIEngine, Segment
from core.utils import FileUtils, VideoUtils, ProjectManager


class MainWindow(ctk.CTk):
    """主窗口类"""

    def __init__(self):
        super().__init__()

        # 初始化配置
        self.config_manager = ConfigManager()
        self.ai_engine = AIEngine(self.config_manager)
        self.project_manager: Optional[ProjectManager] = None

        # 数据
        self.segments: List[Segment] = []
        self.materials = {
            "characters": [],  # 人物素材
            "scenes": [],      # 场景素材
            "voices": []       # 声线素材
        }

        # 未保存状态追踪
        self._has_unsaved_changes = False

        # 安全模式状态（默认为 True，即安全模式，手动提交）
        self._safe_mode = True

        # 设置窗口
        self.title("视频片段拆分工具 v1.0")
        self.geometry("1400x900")
        self.minsize(1200, 700)

        # 设置自定义主题
        ctk.set_appearance_mode("light")
        self._setup_custom_theme()

        # 创建UI
        self._create_ui()

        # 检查API配置
        self._check_api_config()

        # 绑定窗口关闭事件
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _setup_custom_theme(self):
        """配置自定义主题样式"""
        # 使用蓝色基础主题
        ctk.set_default_color_theme("blue")

        # 修改主题配置
        theme = ctk.ThemeManager.theme

        # 全局配置
        theme["CTk"]["fg_color"] = ["#fafafa", "#1a1a1a"]

        # 框架样式
        theme["CTkFrame"]["corner_radius"] = 12
        theme["CTkFrame"]["border_width"] = 1
        theme["CTkFrame"]["fg_color"] = ["#ffffff", "#242424"]
        theme["CTkFrame"]["border_color"] = ["#e0e0e0", "#3a3a3a"]

        # 按钮样式
        theme["CTkButton"]["corner_radius"] = 8
        theme["CTkButton"]["fg_color"] = ["#3b7de8", "#3b7de8"]
        theme["CTkButton"]["hover_color"] = ["#2e64c4", "#2e64c4"]

        # 输入框样式
        theme["CTkEntry"]["corner_radius"] = 8
        theme["CTkEntry"]["border_width"] = 2
        theme["CTkEntry"]["fg_color"] = ["#ffffff", "#2b2b2b"]
        theme["CTkEntry"]["border_color"] = ["#d0d0d0", "#404040"]

        # 文本框样式
        theme["CTkTextbox"]["corner_radius"] = 8
        theme["CTkTextbox"]["border_width"] = 1
        theme["CTkTextbox"]["fg_color"] = ["#ffffff", "#2b2b2b"]
        theme["CTkTextbox"]["border_color"] = ["#d0d0d0", "#404040"]

        # 滚动框架样式
        theme["CTkScrollableFrame"]["corner_radius"] = 8
        theme["CTkScrollableFrame"]["border_width"] = 1
        theme["CTkScrollableFrame"]["fg_color"] = ["#ffffff", "#242424"]
        theme["CTkScrollableFrame"]["border_color"] = ["#e0e0e0", "#3a3a3a"]

        # 设置全局字体
        self.option_add("*Font", ("Microsoft YaHei UI", 12))

    def _add_button_effects(self, button):
        """给按钮添加悬浮动效"""
        original_fg_color = button.cget("fg_color")
        original_hover_color = button.cget("hover_color")

        def on_enter(event):
            button.configure(cursor="hand2")
            # 轻微提升亮度
            if isinstance(original_fg_color, (list, tuple)):
                # 明暗模式都有
                light_color = original_fg_color[0]
                dark_color = original_fg_color[1]
                button.configure(fg_color=[self._lighten_color(light_color, 0.1), self._lighten_color(dark_color, 0.1)])
            else:
                button.configure(fg_color=self._lighten_color(original_fg_color, 0.1))

        def on_leave(event):
            # 恢复原颜色
            button.configure(fg_color=original_fg_color, hover_color=original_hover_color)

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def _lighten_color(self, color, amount=0.1):
        """提亮颜色"""
        if isinstance(color, str) and color.startswith("#"):
            # 十六进制颜色
            r = int(color[1:3], 16)
            g = int(color[3:5], 16)
            b = int(color[5:7], 16)
            r = min(255, int(r + (255 - r) * amount))
            g = min(255, int(g + (255 - g) * amount))
            b = min(255, int(b + (255 - b) * amount))
            return f"#{r:02x}{g:02x}{b:02x}"
        return color

    def _create_ui(self):
        """创建UI组件"""
        # 主容器
        self.main_container = ctk.CTkFrame(self)
        self.main_container.pack(fill="both", expand=True, padx=12, pady=12)

        # 顶部工具栏（玻璃拟态效果）
        self._create_toolbar()

        # 中间内容区（左侧素材面板 + 右侧主内容）
        self.content_frame = ctk.CTkFrame(
            self.main_container,
            fg_color=["#ffffff", "#1e1e1e"]
        )
        self.content_frame.pack(fill="both", expand=True, pady=(10, 0))

        # 左侧素材面板
        self._create_material_panel()

        # 右侧主内容区
        self._create_main_content()

        # 底部状态栏
        self._create_status_bar()

    def _create_toolbar(self):
        """创建工具栏"""
        self.toolbar = ctk.CTkFrame(
            self.main_container,
            height=60,
            fg_color=["#f5f7fa", "#1e1e1e"],
            border_width=0
        )
        self.toolbar.pack(fill="x", pady=(0, 12))
        self.toolbar.pack_propagate(False)

        # 左侧按钮组
        left_btn_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        left_btn_frame.pack(side="left", padx=10)

        # 新建项目按钮
        self.btn_new_project = ctk.CTkButton(
            left_btn_frame, text="📁 新建项目", width=110, height=32,
            command=self._on_new_project,
            fg_color="#27ae60", hover_color="#229954"
        )
        self.btn_new_project.pack(side="left", padx=6)

        # 打开项目按钮
        self.btn_open_project = ctk.CTkButton(
            left_btn_frame, text="📂 打开项目", width=110, height=32,
            command=self._on_open_project,
            fg_color="#3498db", hover_color="#2980b9"
        )
        self.btn_open_project.pack(side="left", padx=6)

        # 设置按钮
        self.btn_settings = ctk.CTkButton(
            left_btn_frame, text="⚙️ 设置", width=90, height=32,
            command=self._on_settings,
            fg_color="#95a5a6", hover_color="#7f8c8d"
        )
        self.btn_settings.pack(side="left", padx=6)

        # 导入脚本按钮（使用OptionMenu实现多格式选择）
        self.btn_import_script = ctk.CTkOptionMenu(
            left_btn_frame,
            values=["📥 导入JSON", "📥 导入Excel"],
            width=120, height=32,
            command=self._on_import_script_menu,
            fg_color="#16a085", button_color="#1abc9c",
            button_hover_color="#16a085", dropdown_fg_color="#16a085"
        )
        self.btn_import_script.set("📥 导入脚本")
        self.btn_import_script.pack(side="left", padx=6)

        # 导出脚本按钮（使用OptionMenu实现多格式选择）
        self.btn_export_script = ctk.CTkOptionMenu(
            left_btn_frame,
            values=["📤 导出JSON", "📤 导出Excel"],
            width=120, height=32,
            command=self._on_export_script_menu,
            fg_color="#2980b9", button_color="#3498db",
            button_hover_color="#2980b9", dropdown_fg_color="#2980b9"
        )
        self.btn_export_script.set("📤 导出脚本")
        self.btn_export_script.pack(side="left", padx=6)

        # 添加按钮动效
        self._add_button_effects(self.btn_new_project)
        self._add_button_effects(self.btn_open_project)
        self._add_button_effects(self.btn_settings)

        # 中间：API状态
        self.api_status_label = ctk.CTkLabel(
            self.toolbar, text="⚠️ 未配置API",
            text_color="orange"
        )
        self.api_status_label.pack(side="left", padx=20)

        # 右侧：安全模式切换开关
        self.safe_mode_switch = ctk.CTkSwitch(
            self.toolbar, text="安全模式",
            command=self._on_toggle_safe_mode,
            switch_width=50, switch_height=24,
            corner_radius=12
        )
        self.safe_mode_switch.select()  # 默认选中安全模式
        self.safe_mode_switch.pack(side="right", padx=6)

        # 右侧：输出目录
        right_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        right_frame.pack(side="right", padx=10)

        self.output_label = ctk.CTkLabel(
            right_frame, text="输出目录: 未选择"
        )
        self.output_label.pack(side="left", padx=5)

        self.btn_select_output = ctk.CTkButton(
            right_frame, text="选择目录", width=90, height=32,
            command=self._on_select_output,
            fg_color="#8e44ad", hover_color="#7d3c98"
        )
        self.btn_select_output.pack(side="left", padx=6)
        self._add_button_effects(self.btn_select_output)

    def _create_material_panel(self):
        """创建左侧素材面板"""
        self.material_panel = MaterialPanel(
            self.content_frame,
            on_material_change=self._on_material_change,
            get_project_dir=self._get_project_dir
        )
        self.material_panel.pack(side="left", fill="y", padx=(0, 10))

    def _get_project_dir(self) -> Optional[str]:
        """获取当前项目目录"""
        if self.project_manager and self.project_manager.project_dir:
            return str(self.project_manager.project_dir)
        return None

    def _get_assets_dir(self) -> Optional[str]:
        """获取当前项目的 assets 目录"""
        if self.project_manager and self.project_manager.project_dir:
            return str(self.project_manager.get_assets_dir())
        return None

    def _create_main_content(self):
        """创建主内容区"""
        self.main_content = ctk.CTkFrame(
            self.content_frame,
            fg_color=["#ffffff", "#1e1e1e"]
        )
        self.main_content.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # 输入区域
        self._create_input_area()

        # 表格区域
        self._create_table_area()

    def _create_input_area(self):
        """创建输入区域"""
        input_frame = ctk.CTkFrame(
            self.main_content,
            fg_color=["#ffffff", "#242424"]
        )
        input_frame.pack(fill="x", pady=(0, 12))

        # 标题
        title_label = ctk.CTkLabel(
            input_frame, text="视频描述/旁白",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title_label.pack(anchor="w", padx=10, pady=(10, 5))

        # 文本输入框
        self.text_input = ctk.CTkTextbox(input_frame, height=120)
        self.text_input.pack(fill="x", padx=12, pady=(0, 12))

        # 绑定文本变化事件（标记为未保存）
        self.text_input.bind("<KeyRelease>", self._on_description_change)

        # 按钮组（已移除灰色边框）
        btn_container = ctk.CTkFrame(input_frame, fg_color="transparent", border_width=0, corner_radius=0)
        btn_container.pack(fill="x", padx=10, pady=(0, 10))

        # 生成按钮
        self.btn_generate = ctk.CTkButton(
            btn_container, text="🎬 拆分片段", width=130, height=34,
            command=self._on_generate,
            fg_color="#e67e22", hover_color="#d35400",
            font=ctk.CTkFont(weight="bold")
        )
        self.btn_generate.pack(side="left", padx=6)

        # 清空按钮
        self.btn_clear = ctk.CTkButton(
            btn_container, text="🗑️ 清空", width=90, height=34,
            command=self._on_clear,
            fg_color="#e74c3c", hover_color="#c0392b"
        )
        self.btn_clear.pack(side="left", padx=6)

        # 保存按钮
        self.btn_save = ctk.CTkButton(
            btn_container, text="💾 保存", width=90, height=34,
            command=self._on_save,
            fg_color="#27ae60", hover_color="#229954"
        )
        self.btn_save.pack(side="left", padx=6)

        # 手动添加片段按钮
        self.btn_add_segment = ctk.CTkButton(
            btn_container, text="➕ 手动添加片段", width=140, height=34,
            command=self._on_manual_add_segment,
            fg_color="#3498db", hover_color="#2980b9"
        )
        self.btn_add_segment.pack(side="left", padx=6)

        # 批量生成视频按钮
        self.btn_batch_generate = ctk.CTkButton(
            btn_container, text="🎬 批量生成视频", width=140, height=34,
            command=self._on_batch_generate_video,
            fg_color="#9b59b6", hover_color="#8e44ad",
            font=ctk.CTkFont(weight="bold")
        )
        self.btn_batch_generate.pack(side="left", padx=6)

        # 添加按钮动效
        self._add_button_effects(self.btn_generate)
        self._add_button_effects(self.btn_clear)
        self._add_button_effects(self.btn_save)
        self._add_button_effects(self.btn_add_segment)
        self._add_button_effects(self.btn_batch_generate)

        # 进度标签
        self.progress_label = ctk.CTkLabel(
            btn_container, text="", text_color="gray"
        )
        self.progress_label.pack(side="right", padx=10)

    def _create_table_area(self):
        """创建表格区域"""
        table_frame = ctk.CTkFrame(
            self.main_content,
            fg_color=["#ffffff", "#242424"]
        )
        table_frame.pack(fill="both", expand=True)

        # 标题
        table_title = ctk.CTkLabel(
            table_frame, text="片段列表",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        table_title.pack(anchor="w", padx=10, pady=10)

        # 表格
        self.segment_table = SegmentTable(
            table_frame,
            on_prompt_edit=self._on_prompt_edit,
            on_video_upload=self._on_video_upload,
            on_generate_grid=self._on_generate_grid,
            on_delete_segment=self._on_delete_segment,
            on_add_segment=self._on_add_segment,
            on_add_reference=self._on_add_reference,
            on_delete_reference=self._on_delete_reference,
            on_generate_video=self._on_generate_video,
            on_batch_generate_video=self._on_batch_generate_video,
            get_assets_dir=self._get_assets_dir,
            on_duration_edit=self._on_duration_edit,
            is_safe_mode=lambda: self._safe_mode  # 传递安全模式获取函数
        )
        self.segment_table.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _create_status_bar(self):
        """创建状态栏"""
        self.status_bar = ctk.CTkFrame(self.main_container, height=30)
        self.status_bar.pack(fill="x", pady=(10, 0))
        self.status_bar.pack_propagate(False)

        self.status_label = ctk.CTkLabel(
            self.status_bar, text="就绪"
        )
        self.status_label.pack(side="left", padx=10)

        self.segment_count_label = ctk.CTkLabel(
            self.status_bar, text="片段: 0"
        )
        self.segment_count_label.pack(side="right", padx=10)

        # 未保存状态标签
        self.unsaved_label = ctk.CTkLabel(
            self.status_bar, text="", text_color="orange"
        )
        self.unsaved_label.pack(side="right", padx=10)

    def _check_api_config(self):
        """检查API配置"""
        api_settings = self.config_manager.get_api_settings()
        if api_settings.get("api_key"):
            self.api_status_label.configure(
                text=f"✅ {api_settings.get('model', '已配置')}",
                text_color="green"
            )
        else:
            self.api_status_label.configure(
                text="⚠️ 未配置API - 请点击设置按钮配置",
                text_color="orange"
            )

    # 事件处理
    def _on_new_project(self):
        """新建项目"""
        # 选择项目目录
        project_dir = filedialog.askdirectory(title="选择项目保存目录")
        if not project_dir:
            return

        # 输入项目名称
        dialog = ctk.CTkInputDialog(
            text="请输入项目名称:",
            title="新建项目"
        )
        project_name = dialog.get_input()
        if not project_name:
            return

        # 创建项目
        self.project_manager = ProjectManager(project_dir)
        self.project_manager.create_project(project_name)

        # 更新输出目录
        self.output_label.configure(
            text=f"输出目录: {self.project_manager.project_dir}"
        )

        # 保存配置
        self.config_manager.set_output_dir(str(self.project_manager.project_dir))

        # 清空界面数据（新项目应该是空的）
        self.text_input.delete("1.0", "end")
        self.segments = []
        self.segment_table.clear()
        self.segment_count_label.configure(text="片段: 0")
        self.material_panel.clear()
        self.materials = {"characters": [], "scenes": [], "voices": []}
        self._set_unsaved(False)

        self._update_status(f"项目已创建: {project_name}")

    def _on_open_project(self):
        """打开已有项目"""
        # 获取上次使用的目录或默认目录
        last_dir = self.config_manager.get_output_dir()

        # 选择项目目录
        project_path = filedialog.askdirectory(
            title="选择项目目录",
            initialdir=last_dir if last_dir else None
        )
        if not project_path:
            return

        # 检查是否是有效的项目目录
        if not os.path.exists(os.path.join(project_path, "segments")):
            # 如果选择的不是项目目录，可能是包含多个项目的父目录
            # 列出其中的项目供用户选择
            temp_manager = ProjectManager(project_path)
            projects = temp_manager.list_projects(project_path)

            if not projects:
                messagebox.showwarning("警告", "所选目录下没有找到有效的项目\n请选择包含 segments 子目录的项目目录")
                return

            # 如果只有一个项目，直接选择
            if len(projects) == 1:
                project_path = projects[0]["path"]
            else:
                # 多个项目，弹出选择对话框
                selected = self._show_project_selection_dialog(projects)
                if not selected:
                    return
                project_path = selected

        # 加载项目
        self.project_manager = ProjectManager(os.path.dirname(project_path))
        if self.project_manager.load_project(project_path):
            # 更新输出目录显示
            self.output_label.configure(
                text=f"输出目录: {self.project_manager.project_dir}"
            )

            # 保存配置
            self.config_manager.set_output_dir(str(self.project_manager.project_dir))

            # 尝试加载项目信息
            project_info = self.project_manager.load_project_info()
            project_name = self.project_manager.current_project
            if project_info and project_info.get("name"):
                project_name = project_info["name"]

            # 先清空所有旧数据，避免新项目数据为空时显示旧数据
            self.segments = []
            self.segment_table.clear()
            self.segment_count_label.configure(text="片段: 0")
            self.text_input.delete("1.0", "end")
            self.material_panel.clear()
            self.materials = {"characters": [], "scenes": [], "voices": []}
            self._set_unsaved(False)

            # 加载所有数据
            loaded_items = []

            # 加载素材元数据
            if self._load_materials():
                total_materials = sum(len(v) for v in self.materials.values())
                loaded_items.append(f"{total_materials} 个素材")

            # 加载视频描述/旁白
            if self._load_description():
                loaded_items.append("视频描述")

            # 加载片段数据
            if self._load_segments():
                loaded_items.append(f"{len(self.segments)} 个片段")

            # 更新状态
            if loaded_items:
                self._update_status(f"已打开项目: {project_name} (已加载: {', '.join(loaded_items)})")
            else:
                self._update_status(f"已打开项目: {project_name}")
        else:
            messagebox.showerror("错误", "无法加载项目")

    def _show_project_selection_dialog(self, projects: List[Dict]) -> Optional[str]:
        """显示项目选择对话框"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("选择项目")
        dialog.geometry("550x450")
        dialog.transient(self)
        dialog.grab_set()

        # 淡入动效
        dialog.attributes("-alpha", 0.0)
        def fade_in():
            current = float(dialog.attributes("-alpha"))
            if current < 1.0:
                dialog.attributes("-alpha", min(current + 0.15, 1.0))
                dialog.after(20, fade_in)

        dialog.after(0, fade_in)

        selected_path = [None]  # 用列表存储以便在闭包中修改

        # 标题
        title_label = ctk.CTkLabel(
            dialog, text="选择要打开的项目",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        title_label.pack(pady=10)

        # 项目列表框架
        list_frame = ctk.CTkScrollableFrame(dialog)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)

        def on_select(project_path: str):
            selected_path[0] = project_path
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        # 显示项目列表
        for project in projects:
            project_frame = ctk.CTkFrame(list_frame)
            project_frame.pack(fill="x", pady=5)

            # 项目名称
            name_label = ctk.CTkLabel(
                project_frame, text=project["name"],
                font=ctk.CTkFont(size=13, weight="bold"),
                anchor="w"
            )
            name_label.pack(fill="x", padx=10, pady=(5, 0))

            # 创建时间
            time_label = ctk.CTkLabel(
                project_frame, text=f"创建时间: {project['create_time']}",
                font=ctk.CTkFont(size=11),
                text_color="gray",
                anchor="w"
            )
            time_label.pack(fill="x", padx=10)

            # 描述（如果有）
            if project.get("description"):
                desc_label = ctk.CTkLabel(
                    project_frame, text=project["description"],
                    font=ctk.CTkFont(size=11),
                    text_color="gray",
                    anchor="w"
                )
                desc_label.pack(fill="x", padx=10)

            # 选择按钮
            select_btn = ctk.CTkButton(
                project_frame, text="选择", width=60,
                command=lambda p=project["path"]: on_select(p)
            )
            select_btn.pack(side="right", padx=10, pady=5)

        # 底部按钮
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)

        cancel_btn = ctk.CTkButton(
            btn_frame, text="取消", width=80,
            command=on_cancel,
            fg_color="gray"
        )
        cancel_btn.pack(side="right")

        # 等待对话框关闭
        self.wait_window(dialog)

        return selected_path[0]

    def _on_settings(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self, self.config_manager)

        # 淡入动效
        dialog.attributes("-alpha", 0.0)
        def fade_in():
            current = float(dialog.attributes("-alpha"))
            if current < 1.0:
                dialog.attributes("-alpha", min(current + 0.15, 1.0))
                dialog.after(20, fade_in)

        dialog.after(0, fade_in)

        dialog.grab_set()
        self.wait_window(dialog)
        self._check_api_config()

    def _on_select_output(self):
        """选择输出目录"""
        output_dir = filedialog.askdirectory(title="选择输出目录")
        if output_dir:
            self.output_label.configure(text=f"输出目录: {output_dir}")
            self.config_manager.set_output_dir(output_dir)

    def _on_toggle_safe_mode(self):
        """切换安全模式"""
        # 获取开关的值
        new_state = self.safe_mode_switch.get()

        if not new_state:
            # 当前是安全模式，切换到非安全模式
            result = messagebox.askyesno(
                "确认切换",
                "非安全模式下视频会自动提交，建议审视视频内容、比例和秒数等选项后再手动提交，以免造成积分浪费。\n\n是否继续切换？"
            )
            if not result:
                # 用户取消，恢复开关状态
                self.safe_mode_switch.set(True)
                return
            self._safe_mode = False
            self._update_status("已切换到非安全模式（自动提交）")
        else:
            # 当前是非安全模式，切换到安全模式
            self._safe_mode = True
            self._update_status("已切换到安全模式（手动提交）")

    def _on_material_change(self, material_type: str, materials: List[Dict]):
        """素材变化回调"""
        self.materials[material_type] = materials
        self._set_unsaved(True)  # 标记为未保存

    def _on_description_change(self, event=None):
        """视频描述/旁白变化回调"""
        self._set_unsaved(True)  # 标记为未保存

    def _on_generate(self):
        """生成片段"""
        # 检查API配置
        api_settings = self.config_manager.get_api_settings()
        if not api_settings.get("api_key"):
            messagebox.showwarning("警告", "请先配置API设置")
            return

        # 获取输入文本
        text = self.text_input.get("1.0", "end-1c").strip()
        if not text:
            messagebox.showwarning("警告", "请输入视频描述或旁白")
            return

        # 禁用按钮
        self.btn_generate.configure(state="disabled", text="处理中...")

        # 在后台线程执行
        def generate_task():
            try:
                success, segments = self.ai_engine.split_into_segments(
                    text,
                    self.materials,
                    on_progress=self._update_progress
                )

                # 在主线程更新UI
                self.after(0, lambda: self._on_generate_complete(success, segments))
            except Exception as e:
                self.after(0, lambda: self._on_generate_error(str(e)))

        thread = threading.Thread(target=generate_task)
        thread.daemon = True
        thread.start()

    def _on_generate_complete(self, success: bool, segments: List[Segment]):
        """生成完成回调"""
        self.btn_generate.configure(state="normal", text="🎬 拆分片段")

        if success and segments:
            self.segments = segments
            self.segment_table.set_segments(segments, self.materials)
            self.segment_count_label.configure(text=f"片段: {len(segments)}")
            self._set_unsaved(True)  # 生成新内容，标记为未保存
            self._update_status(f"已生成 {len(segments)} 个片段")
        else:
            messagebox.showerror("错误", "生成失败，请检查API设置和网络连接")
            self._update_status("生成失败")

    def _on_generate_error(self, error: str):
        """生成错误回调"""
        self.btn_generate.configure(state="normal", text="🎬 拆分片段")
        messagebox.showerror("错误", f"生成失败: {error}")
        self._update_status(f"错误: {error}")

    def _on_clear(self):
        """清空内容"""
        self.text_input.delete("1.0", "end")
        self.segments = []
        self.segment_table.clear()
        self.segment_count_label.configure(text="片段: 0")
        # 清空素材
        self.material_panel.clear()
        self.materials = {"characters": [], "scenes": [], "voices": []}
        self._set_unsaved(False)  # 清空后无未保存状态
        self._update_status("已清空")

    def _on_save(self):
        """保存所有数据（片段、素材、描述）"""
        if not self.project_manager:
            messagebox.showwarning("警告", "请先创建或打开项目")
            return

        # 保存前先同步表格数据
        self._sync_table_data()

        try:
            saved_items = []

            # 1. 保存片段数据
            if self.segments:
                segments_data = []
                for seg in self.segments:
                    segments_data.append({
                        "index": seg.index,
                        "prompt": seg.prompt,
                        "duration": seg.duration,
                        "references": seg.references,
                        "video_path": seg.video_path,
                        "narration": seg.narration
                    })
                self.project_manager.save_segments(segments_data)
                saved_items.append(f"{len(self.segments)} 个片段")

            # 2. 保存素材元数据
            materials_data = self.material_panel.get_materials()
            if any(materials_data.values()):  # 如果有任何素材
                self.project_manager.save_materials(materials_data)
                total_materials = sum(len(v) for v in materials_data.values())
                saved_items.append(f"{total_materials} 个素材")

            # 3. 保存视频描述/旁白
            description = self.text_input.get("1.0", "end-1c").strip()
            if description:
                self.project_manager.save_description(description)
                saved_items.append("视频描述")

            self._set_unsaved(False)  # 清除未保存状态

            if saved_items:
                self._update_status(f"已保存: {', '.join(saved_items)}")
                messagebox.showinfo("保存成功", f"已保存以下内容:\n\n" + "\n".join(f"✓ {item}" for item in saved_items))
            else:
                self._update_status("没有数据需要保存")
                messagebox.showinfo("提示", "没有数据需要保存")

        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {str(e)}")

    def _on_save_silent(self):
        """静默保存所有数据（不弹出提示框）"""
        if not self.project_manager:
            return

        try:
            # 同步表格数据
            self._sync_table_data()

            # 1. 保存片段数据
            if self.segments:
                segments_data = []
                for seg in self.segments:
                    segments_data.append({
                        "index": seg.index,
                        "prompt": seg.prompt,
                        "duration": seg.duration,
                        "references": seg.references,
                        "video_path": seg.video_path,
                        "narration": seg.narration
                    })
                self.project_manager.save_segments(segments_data)

            # 2. 保存素材元数据
            materials_data = self.material_panel.get_materials()
            if any(materials_data.values()):
                self.project_manager.save_materials(materials_data)

            # 3. 保存视频描述/旁白
            description = self.text_input.get("1.0", "end-1c").strip()
            if description:
                self.project_manager.save_description(description)

            self._set_unsaved(False)
            self._update_status(f"已自动保存所有数据")
        except Exception as e:
            print(f"自动保存失败: {e}")

    def _load_segments(self):
        """加载片段数据"""
        if not self.project_manager:
            return False

        try:
            segments_data = self.project_manager.load_segments()
            if segments_data:
                self.segments = []
                for seg_dict in segments_data:
                    segment = Segment(
                        index=seg_dict.get("index", 0),
                        prompt=seg_dict.get("prompt", ""),
                        duration=seg_dict.get("duration", 10.0),
                        references=seg_dict.get("references", []),
                        video_path=seg_dict.get("video_path"),
                        narration=seg_dict.get("narration", "")
                    )
                    self.segments.append(segment)

                self.segment_table.set_segments(self.segments, self.materials)
                self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
                self._set_unsaved(False)  # 加载完成，清除未保存状态
                return True
            return False
        except Exception as e:
            print(f"加载片段数据失败: {e}")
            return False

    def _load_materials(self):
        """加载素材元数据"""
        if not self.project_manager:
            return False

        try:
            materials_data = self.project_manager.load_materials()
            if materials_data:
                # 更新素材面板
                self.material_panel.materials = materials_data
                self.material_panel._switch_tab()

                # 同步到主窗口的 materials 引用
                self.materials = materials_data
                return True
            return False
        except Exception as e:
            print(f"加载素材数据失败: {e}")
            return False

    def _load_description(self):
        """加载视频描述/旁白"""
        if not self.project_manager:
            return False

        try:
            description = self.project_manager.load_description()
            if description:
                self.text_input.delete("1.0", "end")
                self.text_input.insert("1.0", description)
                return True
            return False
        except Exception as e:
            print(f"加载描述数据失败: {e}")
            return False

    def _on_prompt_edit(self, index: int, new_prompt: str):
        """提示词编辑回调"""
        if 0 <= index < len(self.segments):
            self.segments[index].prompt = new_prompt
            self._set_unsaved(True)
            self._update_status(f"已更新片段 {index + 1} 的提示词")

    def _on_duration_edit(self, index: int, new_duration: int):
        """时长编辑回调"""
        if 0 <= index < len(self.segments):
            self.segments[index].duration = float(new_duration)
            self._set_unsaved(True)
            self._update_status(f"已更新片段 {index + 1} 的时长为 {new_duration} 秒")

    def _on_video_upload(self, index: int, video_path: str):
        """视频上传回调"""
        if 0 <= index < len(self.segments):
            self.segments[index].video_path = video_path

            # 如果有项目目录，复制视频到片段目录
            if self.project_manager:
                segment_dir = self.project_manager.get_segment_dir(index + 1)
                dst_path = segment_dir / os.path.basename(video_path)
                FileUtils.copy_file(video_path, str(dst_path))

            self._set_unsaved(True)
            self._update_status(f"已上传视频到片段 {index + 1}")

    def _on_generate_grid(self, index: int):
        """生成九宫格回调"""
        if not 0 <= index < len(self.segments):
            return

        segment = self.segments[index]
        if not segment.video_path:
            messagebox.showwarning("警告", "请先上传视频")
            return

        # 检查是否有项目目录
        if not self.project_manager:
            messagebox.showwarning("警告", "请先创建项目")
            return

        self._update_status(f"正在生成片段 {index + 1} 的九宫格...")
        self.update()

        try:
            # 提取帧（帧图片可以保存在临时目录，生成九宫格后可删除）
            segment_dir = self.project_manager.get_segment_dir(index + 1)
            frames_dir = segment_dir / "frames"
            frames = VideoUtils.extract_frames(
                segment.video_path,
                num_frames=9,
                output_dir=str(frames_dir)
            )

            if len(frames) < 9:
                messagebox.showerror("错误", "提取帧失败")
                return

            # 简化提示词
            success, summary = self.ai_engine.simplify_prompt(segment.prompt, 30)
            if not success:
                summary = f"片段{index + 1}"

            # 生成九宫格，保存到统一的 assets 目录
            assets_dir = self.project_manager.get_assets_dir()
            grid_path = str(assets_dir / f"grid_{index + 1:03d}.png")
            success = VideoUtils.create_nine_grid(
                frames, grid_path, text=summary
            )

            if success:
                self._update_status(f"已生成九宫格: {grid_path}")

                # 添加到相邻片段的引用
                self._add_grid_reference(index, grid_path, summary)

                messagebox.showinfo("成功", f"九宫格已保存到:\n{grid_path}")
            else:
                messagebox.showerror("错误", "生成九宫格失败")

        except Exception as e:
            messagebox.showerror("错误", f"生成失败: {str(e)}")

    def _add_grid_reference(self, index: int, grid_path: str, summary: str):
        """将九宫格引用添加到相邻片段"""
        # 使用文件名引用，所有文件都在 assets 目录下
        grid_filename = os.path.basename(grid_path)

        # 添加到下一个片段
        if index + 1 < len(self.segments):
            next_segment = self.segments[index + 1]
            next_segment.references.append(grid_filename)
            next_segment.prompt += f"\n【前一段剧情纲要及九宫格截图：@[{grid_filename}]】"

        # 更新表格
        self.segment_table.set_segments(self.segments, self.materials)
        self._set_unsaved(True)  # 标记为未保存

    def _on_delete_segment(self, index: int):
        """删除片段"""
        if 0 <= index < len(self.segments):
            del self.segments[index]
            # 重新编号
            for i, seg in enumerate(self.segments):
                seg.index = i + 1
            self.segment_table.set_segments(self.segments, self.materials)
            self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
            self._set_unsaved(True)

    def _on_add_segment(self, after_index: int):
        """添加片段"""
        new_index = after_index + 1
        new_segment = Segment(
            index=new_index,
            prompt="",
            duration=10.0,
            references=[],
            narration=""
        )

        self.segments.insert(new_index, new_segment)

        # 重新编号
        for i, seg in enumerate(self.segments):
            seg.index = i + 1

        self.segment_table.set_segments(self.segments, self.materials)
        self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
        self._set_unsaved(True)

    def _on_add_reference(self, index: int, files: List[str]):
        """添加引用到片段"""
        if 0 <= index < len(self.segments):
            segment = self.segments[index]

            # 添加文件名到引用列表
            for file_name in files:
                if file_name not in segment.references:
                    segment.references.append(file_name)

            # 更新表格显示
            self.segment_table.set_segments(self.segments, self.materials)
            self._set_unsaved(True)
            self._update_status(f"已添加 {len(files)} 个引用到片段 {index + 1}")

    def _on_delete_reference(self, index: int, refs_to_delete: List[str]):
        """从片段删除引用"""
        if 0 <= index < len(self.segments):
            segment = self.segments[index]

            # 从引用列表中删除指定的引用
            deleted_count = 0
            for ref in refs_to_delete:
                if ref in segment.references:
                    segment.references.remove(ref)
                    deleted_count += 1

            # 同时从提示词中移除对应的引用标记 @[文件名]
            for ref in refs_to_delete:
                # 移除 @[文件名] 格式的引用
                ref_pattern = f"@[{ref}]"
                if ref_pattern in segment.prompt:
                    segment.prompt = segment.prompt.replace(ref_pattern, "")
                    # 清理可能残留的空行或多余空格
                    segment.prompt = segment.prompt.strip()

            # 更新表格显示
            self.segment_table.set_segments(self.segments, self.materials)
            self._set_unsaved(True)
            self._update_status(f"已从片段 {index + 1} 删除 {deleted_count} 个引用")

    def _on_generate_video(self, index: int):
        """生成视频（即梦AI）回调"""
        if not 0 <= index < len(self.segments):
            return

        # 关键：先生成前同步表格数据，确保使用最新的提示词
        self._sync_table_data()

        # 再次获取同步后的 segment
        segment = self.segments[index]

        # 检查是否有项目目录
        if not self.project_manager:
            messagebox.showwarning("警告", "请先创建或打开项目")
            return

        # 获取 assets 目录
        assets_dir = str(self.project_manager.get_assets_dir())

        # 导入即梦自动化模块
        try:
            from core.jimeng_automation import generate_jimeng_video_for_segment
        except ImportError as e:
            messagebox.showerror("错误", f"导入即梦模块失败: {e}\n请确保已安装 playwright: pip install playwright")
            return

        # 自动保存（如果有未保存的更改）
        if self._has_unsaved_changes:
            self._on_save_silent()

        # 更新状态
        self._update_status(f"正在为片段 {index + 1} 生成视频...")
        self.update()

        # 定义进度回调
        def on_progress(message: str, step: int = 0, total: int = 0):
            self.after(0, lambda: self._update_status(message))

        # 在后台线程执行
        def generate_task():
            try:
                # 安全模式下不自动提交
                auto_submit = not self._safe_mode
                success, message = generate_jimeng_video_for_segment(
                    segment=segment,
                    assets_dir=assets_dir,
                    on_progress=on_progress,
headless=False,  # 显示浏览器窗口
                    auto_submit=auto_submit  # 传递是否自动提交
                )

                # 在主线程更新UI
                self.after(0, lambda: self._on_generate_video_complete(index, success, message))
            except Exception as e:
                self.after(0, lambda: self._on_generate_video_error(index, str(e)))

        thread = threading.Thread(target=generate_task)
        thread.daemon = True
        thread.start()

    def _on_generate_video_complete(self, index: int, success: bool, message: str):
        """生成视频完成回调"""
        if success:
            self._update_status(f"片段 {index + 1} 视频生成请求已提交")
            messagebox.showinfo("成功", message)
        else:
            self._update_status(f"片段 {index + 1} 视频生成失败")
            messagebox.showerror("错误", message)

    def _on_generate_video_error(self, index: int, error: str):
        """生成视频错误回调"""
        self._update_status(f"片段 {index + 1} 视频生成错误: {error}")
        messagebox.showerror("错误", f"视频生成失败: {error}")

    def _on_manual_add_segment(self):
        """手动添加片段到列表末尾"""
        # 计算新片段的索引
        new_index = len(self.segments) + 1
        new_segment = Segment(
            index=new_index,
            prompt="",
            duration=10.0,
            references=[],
            narration=""
        )

        self.segments.append(new_segment)
        self.segment_table.set_segments(self.segments, self.materials)
        self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
        self._set_unsaved(True)
        self._update_status(f"已手动添加片段 {new_index}")

    def _on_batch_generate_video(self):
        """批量生成视频（即梦AI）"""
        # 获取选中的片段索引
        selected_indices = self.segment_table.get_selected_indices()

        if not selected_indices:
            messagebox.showwarning("提示", "请先勾选要生成视频的片段")
            return

        # 检查是否有项目目录
        if not self.project_manager:
            messagebox.showwarning("警告", "请先创建或打开项目")
            return

        # 同步表格数据
        self._sync_table_data()

        # 获取 assets 目录
        assets_dir = str(self.project_manager.get_assets_dir())

        # 确认对话框
        mode_text = "手动审查后提交" if self._safe_mode else "自动提交"
        result = messagebox.askyesno(
            "批量生成视频",
            f"即将为 {len(selected_indices)} 个片段逐个生成视频\n"
            f"片段序号: {', '.join([str(i+1) for i in selected_indices])}\n\n"
            f"当前模式: {mode_text}\n"
            f"将自动打开浏览器，{mode_text}\n"
            f"是否继续？"
        )
        if not result:
            return

        # 导入即梦自动化模块
        try:
            from core.jimeng_automation import generate_jimeng_video_for_segment
        except ImportError as e:
            messagebox.showerror("错误", f"导入即梦模块失败: {e}\n请确保已安装 playwright: pip install playwright")
            return

        # 自动保存（如果有未保存的更改）
        if self._has_unsaved_changes:
            self._on_save_silent()

        # 禁用批量生成按钮
        self.btn_batch_generate.configure(state="disabled", text="生成中...")

        # 定义进度回调
        def on_progress(message: str, step: int = 0, total: int = 0):
            self.after(0, lambda: self._update_status(message))

        # 在后台线程执行批量生成
        def batch_generate_task():
            success_count = 0
            fail_count = 0
            results = []

            # 安全模式下不自动提交
            auto_submit = not self._safe_mode

            for i, index in enumerate(selected_indices):
                segment = self.segments[index]

                # 更新状态
                self.after(0, lambda idx=index, i=i, total=len(selected_indices):
                    self._update_status(f"正在生成 [{i+1}/{len(selected_indices)}] 片段 {idx+1}..."))

                try:
                    success, message = generate_jimeng_video_for_segment(
                        segment=segment,
                        assets_dir=assets_dir,
                        on_progress=on_progress,
                        headless=False,
                        auto_submit=auto_submit
                    )

                    if success:
                        success_count += 1
                        results.append(f"片段 {index+1}: 成功")
                    else:
                        fail_count += 1
                        results.append(f"片段 {index+1}: 失败 - {message}")

                    # 每个片段之间等待一下，让用户确认
                    if i < len(selected_indices) - 1:
                        time.sleep(2)

                except Exception as e:
                    fail_count += 1
                    results.append(f"片段 {index+1}: 错误 - {str(e)}")

            # 在主线程更新UI
            self.after(0, lambda: self._on_batch_generate_complete(success_count, fail_count, results))

        thread = threading.Thread(target=batch_generate_task)
        thread.daemon = True
        thread.start()

    def _on_batch_generate_complete(self, success_count: int, fail_count: int, results: List[str]):
        """批量生成视频完成回调"""
        self.btn_batch_generate.configure(state="normal", text="🎬 批量生成视频")

        # 清空选择
        self.segment_table.clear_selection()

        # 显示结果
        result_text = f"批量生成完成\n成功: {success_count}, 失败: {fail_count}\n\n"
        result_text += "\n".join(results)

        self._update_status(f"批量生成完成: 成功 {success_count}, 失败 {fail_count}")

        # 关闭全局浏览器
        try:
            from core.jimeng_automation import JimengVideoAutomation
            JimengVideoAutomation.close_global_browser()
        except Exception as e:
            print(f"关闭浏览器失败: {e}")

        if fail_count > 0:
            messagebox.showwarning("批量生成结果", result_text)
        else:
            messagebox.showinfo("批量生成结果", result_text)

    def _on_closing(self):
        """窗口关闭事件处理"""
        if self._has_unsaved_changes:
            # 弹出保存确认对话框
            result = messagebox.askyesnocancel(
                "保存确认",
                "您有未保存的更改，是否保存后再关闭？\n\n"
                "选择【是】保存并关闭\n"
                "选择【否】直接关闭（丢失更改）\n"
                "选择【取消】返回继续编辑"
            )
            if result is None:  # 取消
                return
            elif result:  # 是 - 保存
                self._on_save()
        # 关闭全局浏览器
        try:
            from core.jimeng_automation import JimengVideoAutomation
            JimengVideoAutomation.close_global_browser()
        except Exception as e:
            print(f"关闭浏览器失败: {e}")
        # 关闭窗口
        self.destroy()

    def _update_progress(self, message: str):
        """更新进度"""
        self.after(0, lambda: self.progress_label.configure(text=message))

    def _update_status(self, message: str):
        """更新状态栏"""
        self.status_label.configure(text=message)

    def _set_unsaved(self, unsaved: bool = True):
        """设置未保存状态"""
        self._has_unsaved_changes = unsaved
        if unsaved:
            self.unsaved_label.configure(text="⚠️ 未保存")
            # 更新窗口标题
            self.title("视频片段拆分工具 v1.0 *")
        else:
            self.unsaved_label.configure(text="")
            self.title("视频片段拆分工具 v1.0")

    def _sync_table_data(self):
        """
        同步表格数据到内存
        强制从表格组件获取最新的编辑数据
        """
        for i, row_widgets in enumerate(self.segment_table.row_widgets):
            if i < len(self.segments):
                # 获取提示词文本框中的最新内容
                prompt_text = row_widgets.get("prompt")
                if prompt_text:
                    new_prompt = prompt_text.get("1.0", "end-1c")
                    if self.segments[i].prompt != new_prompt:
                        self.segments[i].prompt = new_prompt
                        self._set_unsaved(True)

    def _on_export_script(self):
        """导出片段脚本为JSON格式"""
        # 同步表格数据
        self._sync_table_data()

        # 选择保存路径
        file_path = filedialog.asksaveasfilename(
            title="导出片段脚本",
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
            initialfile="segments_script.json"
        )

        if not file_path:
            return

        try:
            # 构建导出数据
            export_data = {
                "version": "1.0",
                "export_time": time.strftime("%Y-%m-%d %H:%M:%S"),
                "description": self.text_input.get("1.0", "end-1c").strip(),
                "segments": []
            }

            if self.segments:
                for seg in self.segments:
                    export_data["segments"].append({
                        "index": seg.index,
                        "prompt": seg.prompt,
                        "duration": seg.duration,
                        "references": seg.references,
                        "narration": seg.narration
                    })

            # 写入文件
            import json
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            if self.segments:
                self._update_status(f"已导出 {len(self.segments)} 个片段到: {file_path}")
                messagebox.showinfo("导出成功", f"片段脚本已导出到:\n{file_path}\n\n共 {len(self.segments)} 个片段")
            else:
                self._update_status(f"已导出模板到: {file_path}")
                messagebox.showinfo("导出成功", f"模板已导出到:\n{file_path}\n\n可以编辑此文件后导入使用")

        except Exception as e:
            messagebox.showerror("导出失败", f"导出失败: {str(e)}")

    def _on_export_script_excel(self):
        """导出片段脚本为Excel格式"""
        # 同步表格数据
        self._sync_table_data()

        # 选择保存路径
        file_path = filedialog.asksaveasfilename(
            title="导出片段脚本",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="segments_script.xlsx"
        )

        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "片段脚本"

            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=12, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            wrap_alignment = Alignment(wrap_text=True, vertical='top')

            # 写入表头
            headers = ["序号", "时长(秒)", "提示词", "旁白", "引用文件"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            ws.column_dimensions['A'].width = 8   # 序号
            ws.column_dimensions['B'].width = 10  # 时长
            ws.column_dimensions['C'].width = 60  # 提示词
            ws.column_dimensions['D'].width = 40  # 旁白
            ws.column_dimensions['E'].width = 30  # 引用文件

            if self.segments:
                # 写入数据
                for row, seg in enumerate(self.segments, 2):
                    ws.cell(row=row, column=1, value=seg.index).border = thin_border
                    ws.cell(row=row, column=2, value=seg.duration).border = thin_border

                    prompt_cell = ws.cell(row=row, column=3, value=seg.prompt)
                    prompt_cell.border = thin_border
                    prompt_cell.alignment = wrap_alignment

                    narration_cell = ws.cell(row=row, column=4, value=seg.narration or "")
                    narration_cell.border = thin_border
                    narration_cell.alignment = wrap_alignment

                    refs_cell = ws.cell(row=row, column=5, value=", ".join(seg.references) if seg.references else "")
                    refs_cell.border = thin_border
                    refs_cell.alignment = wrap_alignment

                    # 设置行高
                    ws.row_dimensions[row].height = 60
            else:
                # 空模板：添加示例行说明
                example_row = 2
                ws.cell(row=example_row, column=1, value=1).border = thin_border
                ws.cell(row=example_row, column=2, value=10).border = thin_border
                example_cell = ws.cell(row=example_row, column=3, value="在此输入片段提示词...")
                example_cell.border = thin_border
                example_cell.alignment = wrap_alignment
                example_cell.font = Font(italic=True, color="808080")

                narration_cell = ws.cell(row=example_row, column=4, value="可选：旁白文本")
                narration_cell.border = thin_border
                narration_cell.alignment = wrap_alignment
                narration_cell.font = Font(italic=True, color="808080")

                refs_cell = ws.cell(row=example_row, column=5, value="可选：ref1.png, ref2.jpg")
                refs_cell.border = thin_border
                refs_cell.alignment = wrap_alignment
                refs_cell.font = Font(italic=True, color="808080")

                ws.row_dimensions[example_row].height = 60

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 保存文件
            wb.save(file_path)

            if self.segments:
                self._update_status(f"已导出 {len(self.segments)} 个片段到: {file_path}")
                messagebox.showinfo("导出成功", f"片段脚本已导出到:\n{file_path}\n\n共 {len(self.segments)} 个片段")
            else:
                self._update_status(f"已导出模板到: {file_path}")
                messagebox.showinfo("导出成功", f"模板已导出到:\n{file_path}\n\n可以编辑此文件后导入使用")

        except ImportError:
            messagebox.showerror("导出失败", "缺少 openpyxl 库，请运行: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出失败: {str(e)}")

    def _on_export_script_menu(self, choice: str):
        """导出脚本菜单回调"""
        if "JSON" in choice:
            self._on_export_script()
        elif "Excel" in choice:
            self._on_export_script_excel()

    def _on_import_script_menu(self, choice: str):
        """导入脚本菜单回调"""
        if "JSON" in choice:
            self._on_import_script()
        elif "Excel" in choice:
            self._on_import_script_excel()

    def _on_import_script(self):
        """导入片段脚本(JSON格式)"""
        # 选择文件
        file_path = filedialog.askopenfilename(
            title="导入片段脚本",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            import json
            with open(file_path, 'r', encoding='utf-8') as f:
                import_data = json.load(f)

            # 验证数据格式
            if not isinstance(import_data, dict):
                messagebox.showerror("导入失败", "无效的脚本格式")
                return

            segments_data = import_data.get("segments", [])
            if not segments_data:
                messagebox.showwarning("警告", "脚本中没有片段数据")
                return

            # 显示预览对话框
            if not self._show_import_preview(import_data, file_path):
                return

            # 清空现有数据
            self.segments = []
            self.segment_table.clear()

            # 导入描述
            description = import_data.get("description", "")
            if description:
                self.text_input.delete("1.0", "end")
                self.text_input.insert("1.0", description)

            # 导入片段
            for seg_dict in segments_data:
                segment = Segment(
                    index=seg_dict.get("index", 0),
                    prompt=seg_dict.get("prompt", ""),
                    duration=seg_dict.get("duration", 10.0),
                    references=seg_dict.get("references", []),
                    video_path=seg_dict.get("video_path"),
                    narration=seg_dict.get("narration", "")
                )
                self.segments.append(segment)

            # 重新编号
            for i, seg in enumerate(self.segments):
                seg.index = i + 1

            # 刷新表格
            self.segment_table.set_segments(self.segments, self.materials)
            self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
            self._set_unsaved(True)

            self._update_status(f"已导入 {len(self.segments)} 个片段")
            messagebox.showinfo("导入成功", f"成功导入 {len(self.segments)} 个片段")

        except json.JSONDecodeError as e:
            messagebox.showerror("导入失败", f"JSON解析失败: {str(e)}")
        except Exception as e:
            messagebox.showerror("导入失败", f"导入失败: {str(e)}")

    def _on_import_script_excel(self):
        """导入片段脚本(Excel格式)"""
        # 选择文件
        file_path = filedialog.askopenfilename(
            title="导入片段脚本",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            # 加载Excel文件
            wb = load_workbook(file_path)
            ws = wb.active

            # 获取表头
            headers = [cell.value for cell in ws[1]]
            if not headers:
                messagebox.showerror("导入失败", "Excel文件格式错误：缺少表头")
                return

            # 查找列索引
            col_map = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    header_lower = str(header).lower()
                    if "序号" in header_lower or "index" in header_lower:
                        col_map["index"] = col_idx
                    elif "时长" in header_lower or "duration" in header_lower:
                        col_map["duration"] = col_idx
                    elif "提示词" in header_lower or "prompt" in header_lower or "描述" in header_lower:
                        col_map["prompt"] = col_idx
                    elif "旁白" in header_lower or "narration" in header_lower:
                        col_map["narration"] = col_idx
                    elif "引用" in header_lower or "reference" in header_lower:
                        col_map["references"] = col_idx

            # 至少需要提示词列
            if "prompt" not in col_map:
                messagebox.showerror("导入失败", "Excel文件缺少「提示词」列")
                return

            # 读取数据行
            segments_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 跳过空行
                if not any(row):
                    continue

                seg = {}
                if "index" in col_map and row[col_map["index"] - 1]:
                    seg["index"] = int(row[col_map["index"] - 1])
                if "duration" in col_map and row[col_map["duration"] - 1]:
                    seg["duration"] = float(row[col_map["duration"] - 1])
                if "prompt" in col_map and row[col_map["prompt"] - 1]:
                    seg["prompt"] = str(row[col_map["prompt"] - 1])
                if "narration" in col_map and row[col_map["narration"] - 1]:
                    seg["narration"] = str(row[col_map["narration"] - 1])
                if "references" in col_map and row[col_map["references"] - 1]:
                    # 引用文件可能是逗号分隔的字符串
                    refs_str = str(row[col_map["references"] - 1])
                    seg["references"] = [r.strip() for r in refs_str.split(",") if r.strip()]

                # 至少要有提示词
                if seg.get("prompt"):
                    segments_data.append(seg)

            if not segments_data:
                messagebox.showwarning("警告", "Excel文件中没有有效的片段数据")
                return

            # 构建导入数据用于预览
            import_data = {
                "version": "Excel导入",
                "export_time": "",
                "description": "",
                "segments": segments_data
            }

            # 显示预览对话框
            if not self._show_import_preview(import_data, file_path):
                return

            # 清空现有数据
            self.segments = []
            self.segment_table.clear()

            # 导入片段
            for seg_dict in segments_data:
                segment = Segment(
                    index=seg_dict.get("index", 0),
                    prompt=seg_dict.get("prompt", ""),
                    duration=seg_dict.get("duration", 10.0),
                    references=seg_dict.get("references", []),
                    narration=seg_dict.get("narration", "")
                )
                self.segments.append(segment)

            # 重新编号
            for i, seg in enumerate(self.segments):
                seg.index = i + 1

            # 刷新表格
            self.segment_table.set_segments(self.segments, self.materials)
            self.segment_count_label.configure(text=f"片段: {len(self.segments)}")
            self._set_unsaved(True)

            self._update_status(f"已导入 {len(self.segments)} 个片段")
            messagebox.showinfo("导入成功", f"成功导入 {len(self.segments)} 个片段")

        except ImportError:
            messagebox.showerror("导入失败", "缺少 openpyxl 库，请运行: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("导入失败", f"导入失败: {str(e)}")

    def _show_import_preview(self, import_data: Dict, file_path: str) -> bool:
        """显示导入预览对话框"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("导入预览")
        dialog.geometry("700x500")
        dialog.transient(self)
        dialog.grab_set()

        # 淡入动效
        dialog.attributes("-alpha", 0.0)
        def fade_in():
            current = float(dialog.attributes("-alpha"))
            if current < 1.0:
                dialog.attributes("-alpha", min(current + 0.15, 1.0))
                dialog.after(20, fade_in)
        dialog.after(0, fade_in)

        result = [False]  # 用列表存储以便在闭包中修改

        # 标题
        title_label = ctk.CTkLabel(
            dialog, text="片段脚本预览",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        title_label.pack(pady=10)

        # 文件信息
        info_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        info_frame.pack(fill="x", padx=20)

        info_text = f"文件: {os.path.basename(file_path)}"
        if import_data.get("export_time"):
            info_text += f"  |  导出时间: {import_data['export_time']}"
        if import_data.get("version"):
            info_text += f"  |  版本: {import_data['version']}"

        info_label = ctk.CTkLabel(info_frame, text=info_text, text_color="gray")
        info_label.pack(anchor="w")

        # 描述（如果有）
        description = import_data.get("description", "")
        if description:
            desc_frame = ctk.CTkFrame(dialog)
            desc_frame.pack(fill="x", padx=20, pady=5)
            desc_title = ctk.CTkLabel(desc_frame, text="视频描述:", font=ctk.CTkFont(weight="bold"))
            desc_title.pack(anchor="w", padx=10, pady=(5, 0))
            desc_text = ctk.CTkLabel(desc_frame, text=description[:200] + ("..." if len(description) > 200 else ""),
                                     wraplength=650, justify="left")
            desc_text.pack(anchor="w", padx=10, pady=(0, 5))

        # 片段列表
        segments_data = import_data.get("segments", [])
        list_label = ctk.CTkLabel(dialog, text=f"片段列表 (共 {len(segments_data)} 个):",
                                  font=ctk.CTkFont(weight="bold"))
        list_label.pack(anchor="w", padx=20, pady=(10, 5))

        # 创建可滚动区域
        scroll_frame = ctk.CTkScrollableFrame(dialog)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=5)

        for seg in segments_data[:20]:  # 最多显示20个预览
            seg_frame = ctk.CTkFrame(scroll_frame)
            seg_frame.pack(fill="x", pady=2)

            # 序号和时长
            index_text = f"#{seg.get('index', '?')}"
            duration_text = f"{seg.get('duration', 10)}s"
            header_label = ctk.CTkLabel(
                seg_frame,
                text=f"{index_text} [{duration_text}]",
                font=ctk.CTkFont(weight="bold"),
                width=80
            )
            header_label.pack(side="left", padx=5)

            # 提示词预览
            prompt = seg.get("prompt", "")
            prompt_preview = prompt[:100] + "..." if len(prompt) > 100 else prompt
            prompt_label = ctk.CTkLabel(
                seg_frame,
                text=prompt_preview if prompt_preview else "(空)",
                text_color="gray" if not prompt_preview else None
            )
            prompt_label.pack(side="left", padx=5, fill="x", expand=True)

        if len(segments_data) > 20:
            more_label = ctk.CTkLabel(scroll_frame, text=f"... 还有 {len(segments_data) - 20} 个片段",
                                      text_color="gray")
            more_label.pack(pady=5)

        # 底部按钮
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)

        def on_confirm():
            result[0] = True
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        cancel_btn = ctk.CTkButton(
            btn_frame, text="取消", width=80,
            command=on_cancel,
            fg_color="gray"
        )
        cancel_btn.pack(side="right", padx=5)

        confirm_btn = ctk.CTkButton(
            btn_frame, text="导入", width=80,
            command=on_confirm
        )
        confirm_btn.pack(side="right", padx=5)

        # 等待对话框关闭
        self.wait_window(dialog)

        return result[0]